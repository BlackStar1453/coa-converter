#!/usr/bin/env python3
"""
XLSX Template Filler
基于 TemplateLayout 动态填充 XLSX 模板，替代硬编码的单元格引用。
"""

import os
import re
import sys
import shutil
import logging
import zipfile
import tempfile
from typing import Optional, Dict, List, Tuple
import openpyxl

from template_detector import TemplateLayout, _col_letter
from coa_converter import COAData, normalize_item_name, convert_date

logger = logging.getLogger(__name__)


def fill_xlsx(coa: COAData, layout: TemplateLayout, template_path: str, output_path: str):
    """根据 layout 动态填充 XLSX 模板"""
    logger.info(f'[XLSX填充] 模板类型: {layout.template_type}')
    logger.info(f'[XLSX填充] 模板: {template_path} → {output_path}')

    shutil.copy2(template_path, output_path)
    if sys.platform != 'win32':
        os.chmod(output_path, 0o644)

    # Flow Chart 使用直接 ZIP/XML 操作以保留 SmartArt 图表数据
    if layout.template_type == 'flowchart':
        _fill_flowchart_xml(coa, layout, output_path)
        logger.info(f'[XLSX填充] 文件已保存（XML直接修改）: {output_path}')
        return

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    if layout.template_type in ('coa_assay', 'coa_powder', 'coa_ratio'):
        _fill_coa(ws, coa, layout)
    elif layout.template_type == 'allergen':
        _fill_allergen(ws, coa, layout)
    else:
        logger.warning(f'[XLSX填充] 未知模板类型: {layout.template_type}，尝试按 COA 方式填充')
        _fill_coa(ws, coa, layout)

    wb.save(output_path)
    logger.info(f'[XLSX填充] 文件已保存: {output_path}')


# ============ COA 模板填充 ============

def _fill_coa(ws, coa: COAData, layout: TemplateLayout):
    """填充 COA 类型模板（Assay/Powder/Ratio）"""
    _fill_coa_header(ws, coa, layout)
    _fill_coa_assay(ws, coa, layout)
    _fill_coa_data_rows(ws, coa, layout)
    _fill_coa_packing(ws, coa, layout)


def _fill_coa_header(ws, coa: COAData, layout: TemplateLayout):
    """填充头部字段"""
    for std_key, mapping in layout.header_fields.items():
        value = coa.header.get(std_key, "")
        if not value:
            # 清空模板预填值
            cell_ref = mapping.value_pos.cell_ref
            if cell_ref:
                ws[cell_ref] = ""
            else:
                ws.cell(row=mapping.value_pos.row, column=mapping.value_pos.col, value="")
            if std_key in ('product_name', 'batch_number', 'mfg_date', 'exp_date'):
                logger.warning(f'[XLSX填充-头部] 缺少字段: {std_key}，已清空模板默认值')
            continue

        if std_key in ("mfg_date", "exp_date"):
            value = convert_date(value)

        cell_ref = mapping.value_pos.cell_ref
        if cell_ref:
            ws[cell_ref] = value
            logger.info(f'[XLSX填充-头部] {cell_ref} = {value}')
        else:
            r, c = mapping.value_pos.row, mapping.value_pos.col
            ws.cell(row=r, column=c, value=value)
            logger.info(f'[XLSX填充-头部] ({r},{c}) = {value}')


def _fill_coa_assay(ws, coa: COAData, layout: TemplateLayout):
    """填充 Assay/Ratio 行"""
    assay_row = layout.table_rows.get('assay')
    if not assay_row:
        logger.info('[XLSX填充-Assay] 模板中无 Assay/Ratio 行（Powder 模板）')
        return

    spec_col = layout.data_columns.get('spec', 3)
    result_col = layout.data_columns.get('result', 5)
    method_col = layout.data_columns.get('method', 6)

    if not coa.assay:
        logger.warning('[XLSX填充-Assay] PDF 中未找到 Assay 数据，清空模板值')
        ws.cell(row=assay_row, column=spec_col, value="")
        ws.cell(row=assay_row, column=result_col, value="")
        ws.cell(row=assay_row, column=method_col, value="")
        return

    spec_val = coa.assay.get("specification", "")
    result_val = coa.assay.get("result", "")
    method_val = coa.assay.get("method", "")

    ws.cell(row=assay_row, column=spec_col, value=spec_val)
    ws.cell(row=assay_row, column=result_col, value=result_val)
    ws.cell(row=assay_row, column=method_col, value=method_val)
    logger.info(f'[XLSX填充-Assay] Row {assay_row}: spec={spec_val}, result={result_val}, method={method_val}')


def _fill_coa_data_rows(ws, coa: COAData, layout: TemplateLayout):
    """填充 Analytical Data + Microbiology 数据行"""
    spec_col = layout.data_columns.get('spec', 3)
    result_col = layout.data_columns.get('result', 5)
    method_col = layout.data_columns.get('method', 6)

    # 构建已提取数据索引
    extracted = {}
    for item in coa.analytical_items + coa.microbiology_items:
        key = normalize_item_name(item["item"])
        if key:
            extracted[key] = item

    # 遍历模板中检测到的所有数据行
    for item_key, row_num in layout.table_rows.items():
        if item_key == 'assay':
            continue  # 已在 _fill_coa_assay 中处理

        if item_key in extracted:
            item = extracted[item_key]
            spec_val = item.get("specification", "")
            result_val = item.get("result", "")
            method_val = item.get("method", "")
            ws.cell(row=row_num, column=spec_col, value=spec_val)
            ws.cell(row=row_num, column=result_col, value=result_val)
            ws.cell(row=row_num, column=method_col, value=method_val)
            logger.info(f'[XLSX填充-数据] Row {row_num} ({item_key}): '
                        f'spec={spec_val}, result={result_val}')
        else:
            # PDF 中无此检测项 → 清空模板预填值
            ws.cell(row=row_num, column=spec_col, value="")
            ws.cell(row=row_num, column=result_col, value="")
            ws.cell(row=row_num, column=method_col, value="")
            logger.warning(f'[XLSX填充-数据] Row {row_num} ({item_key}): PDF 中未找到对应数据，已清空模板默认值')


def _fill_coa_packing(ws, coa: COAData, layout: TemplateLayout):
    """填充 Packing & Storage"""
    if not coa.packing_storage:
        # PDF 中无包装信息 → 清空模板预填值
        if layout.packing_position:
            ref = layout.packing_position.cell_ref
            if ref:
                ws[ref] = ""
            else:
                ws.cell(row=layout.packing_position.row,
                        column=layout.packing_position.col, value="")
        if layout.storage_position:
            ref = layout.storage_position.cell_ref
            if ref:
                ws[ref] = ""
            else:
                ws.cell(row=layout.storage_position.row,
                        column=layout.storage_position.col, value="")
        logger.info('[XLSX填充-包装] PDF 无包装信息，已清空模板默认值')
        return

    parts = re.split(r'(?:store\s)', coa.packing_storage, maxsplit=1, flags=re.IGNORECASE)

    if layout.packing_position:
        packing_text = parts[0].strip().rstrip('.') if len(parts) >= 1 else ""
        ref = layout.packing_position.cell_ref
        if ref:
            ws[ref] = packing_text
        else:
            ws.cell(row=layout.packing_position.row,
                    column=layout.packing_position.col,
                    value=packing_text)
        if packing_text:
            logger.info(f'[XLSX填充-包装] {ref or layout.packing_position.row} = {packing_text[:50]}...')
        else:
            logger.info(f'[XLSX填充-包装] {ref or layout.packing_position.row} = (空，已清除模板默认值)')

    if layout.storage_position and len(parts) >= 2:
        storage_text = "Store " + parts[1].strip()
        ref = layout.storage_position.cell_ref
        if ref:
            ws[ref] = storage_text
        else:
            ws.cell(row=layout.storage_position.row,
                    column=layout.storage_position.col,
                    value=storage_text)
        logger.info(f'[XLSX填充-存储] {ref or layout.storage_position.row} = {storage_text[:50]}...')


# ============ Allergen 模板填充 ============

def _fill_allergen(ws, coa: COAData, layout: TemplateLayout):
    """填充 Allergen 模板 — 主要替换产品名"""
    product_name = coa.header.get('product_name', '')

    if product_name:
        pm = layout.header_fields.get('product_name')
        if pm and pm.value_pos and pm.value_pos.cell_ref:
            # "Product Name: " 后面追加产品名
            current = ws[pm.value_pos.cell_ref].value or ''
            if current.endswith(': ') or current.endswith(':'):
                ws[pm.value_pos.cell_ref] = current + product_name
            else:
                ws[pm.value_pos.cell_ref] = f'Product Name: {product_name}'
            logger.info(f'[XLSX填充-Allergen] 产品名: {product_name}')

    # Allergen 的检测项数据通常不来自 COA PDF，所以不填数据列
    logger.info('[XLSX填充-Allergen] Allergen 模板仅填充产品名，数据列保持模板默认值')


# ============ Flow Chart 模板填充（XML 直接操作，保留 SmartArt） ============

def _fill_flowchart_xml(coa: COAData, layout: TemplateLayout, output_path: str):
    """使用直接 ZIP/XML 操作填充 Flow Chart 标题，保留 SmartArt 图表数据。
    openpyxl 不支持 SmartArt，load/save 会丢弃 xl/diagrams/ 内容。
    修改 xl/sharedStrings.xml 中的共享字符串，避免重新序列化 sheet1.xml 导致命名空间损坏。
    """
    product_name = coa.header.get('product_name', '')
    if not product_name:
        logger.info('[XLSX填充-FlowChart] 产品名为空，跳过填充')
        return

    title_text = f'{product_name} Flow Chart'

    # 模板有多个工作表（生粉/汁粉/冻干粉/Extract/浓缩汁），
    # 需要根据产品名选择正确的工作表，替换其 A2 标题，并设为活动工作表。
    SHARED_STRINGS_PATH = 'xl/sharedStrings.xml'
    WORKBOOK_PATH = 'xl/workbook.xml'
    WORKBOOK_RELS_PATH = 'xl/_rels/workbook.xml.rels'
    A2_INDEX_PATTERN = re.compile(r'<c\s+r="A2"[^>]*t="s"[^>]*>\s*<v>(\d+)</v>', re.DOTALL)
    SI_PATTERN = re.compile(r'<si>.*?</si>', re.DOTALL)

    # 根据产品名选择工作表（匹配 sheet name 中的关键词）
    SHEET_KEYWORDS = [
        ('extract', '所有Extract'),
        ('juice concentrate', '浓缩汁'),
        ('concentrate', '浓缩汁'),
        ('freeze dried', '冻干粉'),
        ('freeze-dried', '冻干粉'),
        ('juice powder', '汁粉'),
    ]

    def _select_sheet_tab(product: str, sheet_names: list) -> int:
        """根据产品名返回应激活的工作表 tab 索引"""
        pl = product.lower()
        for keyword, sheet_hint in SHEET_KEYWORDS:
            if keyword in pl:
                for idx, sn in enumerate(sheet_names):
                    if sheet_hint in sn:
                        return idx
        return 0  # 默认: 所有生粉

    tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(tmp_fd)
    try:
        with zipfile.ZipFile(output_path, 'r') as zin:
            # 1. 解析 workbook.xml 获取工作表名称和顺序
            wb_xml = zin.read(WORKBOOK_PATH).decode('UTF-8')
            sheet_entries = re.findall(
                r'<sheet\s+name="([^"]+)"\s+sheetId="\d+"\s+r:id="(rId\d+)"', wb_xml)
            sheet_names = [name for name, _ in sheet_entries]
            sheet_rids = {name: rid for name, rid in sheet_entries}

            # 2. 解析 rels 获取 rId → sheet 文件名映射
            rels_xml = zin.read(WORKBOOK_RELS_PATH).decode('UTF-8')
            rid_to_file = {}
            for rid, target in re.findall(r'<Relationship\s+Id="(rId\d+)"[^>]*Target="([^"]+)"', rels_xml):
                rid_to_file[rid] = target

            # 3. 选择目标工作表
            target_tab = _select_sheet_tab(product_name, sheet_names)
            target_sheet_name = sheet_names[target_tab]
            target_rid = sheet_rids[target_sheet_name]
            target_file = 'xl/' + rid_to_file[target_rid]
            logger.info(f'[XLSX填充-FlowChart] 选择工作表: tab={target_tab}, '
                        f'name="{target_sheet_name}", file={target_file}')

            # 4. 从目标工作表的 XML 中获取 A2 的共享字符串索引
            sheet_xml = zin.read(target_file).decode('UTF-8')
            m = A2_INDEX_PATTERN.search(sheet_xml)
            if not m:
                logger.warning(f'[XLSX填充-FlowChart] 未在 {target_file} 中找到 A2 共享字符串引用')
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
                return
            target_idx = int(m.group(1))
            logger.info(f'[XLSX填充-FlowChart] A2 引用共享字符串索引: {target_idx}')

        # 5. 构建新的 rich text <si>（产品名红色 + " Flow Chart" 黑色）
        NEW_SI = (
            '<si><r><rPr><b/><sz val="18"/><color rgb="FFFF0000"/>'
            '<rFont val="Times New Roman"/><family val="1"/></rPr>'
            f'<t>{product_name}</t></r><r><rPr><b/><sz val="18"/>'
            '<color theme="1"/><rFont val="Times New Roman"/><family val="1"/></rPr>'
            '<t xml:space="preserve"> Flow Chart</t></r></si>'
        )

        with zipfile.ZipFile(output_path, 'r') as zin, \
             zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == SHARED_STRINGS_PATH:
                    # 替换目标共享字符串
                    xml_str = data.decode('UTF-8')
                    si_elements = list(SI_PATTERN.finditer(xml_str))
                    if target_idx < len(si_elements):
                        old_si = si_elements[target_idx]
                        xml_str = xml_str[:old_si.start()] + NEW_SI + xml_str[old_si.end():]
                        logger.info(f'[XLSX填充-FlowChart] sharedStrings 索引 {target_idx} 已替换为: {title_text}')
                    else:
                        logger.warning(f'[XLSX填充-FlowChart] 共享字符串索引 {target_idx} 超出范围 (共 {len(si_elements)} 个)')
                    data = xml_str.encode('UTF-8')
                elif item.filename == WORKBOOK_PATH:
                    # 设置活动工作表
                    xml_str = data.decode('UTF-8')
                    if 'activeTab=' in xml_str:
                        xml_str = re.sub(r'activeTab="\d+"', f'activeTab="{target_tab}"', xml_str)
                    else:
                        xml_str = xml_str.replace('<workbookView ', f'<workbookView activeTab="{target_tab}" ')
                    data = xml_str.encode('UTF-8')
                    logger.info(f'[XLSX填充-FlowChart] 活动工作表设为 tab {target_tab}: "{target_sheet_name}"')
                zout.writestr(item, data)
        shutil.move(tmp_path, output_path)
        logger.info(f'[XLSX填充-FlowChart] 标题: {title_text}（SmartArt 已保留）')
    except Exception as e:
        logger.error(f'[XLSX填充-FlowChart] XML 直接修改失败: {e}')
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise


# ============ 输出验证 ============

def verify_xlsx_output(coa: COAData, layout: TemplateLayout, output_path: str,
                       template_path: Optional[str] = None) -> Dict:
    """
    验证 XLSX 输出文件的数据正确性。
    双重验证：1) 对比 COAData（PDF ground truth）  2) 对比原模板检测"默认值残留"

    Returns:
        {
            "total": int,          # 总验证字段数
            "passed": int,         # 通过的字段数
            "failed": int,         # 失败的字段数
            "accuracy": float,     # 正确率 (0.0-1.0)
            "details": [...]       # 每个字段的验证明细
            "template_defaults_retained": [...]  # 保留了模板默认值的字段
        }
    """
    if layout.template_type not in ('coa_assay', 'coa_powder', 'coa_ratio'):
        logger.info(f'[验证] 跳过非 COA 模板: {layout.template_type}')
        return {"total": 0, "passed": 0, "failed": 0, "accuracy": 1.0,
                "details": [], "template_defaults_retained": []}

    wb = openpyxl.load_workbook(output_path, data_only=True)
    ws = wb.active

    # 加载原模板用于对比检测"默认值残留"
    template_cells = {}
    if template_path:
        try:
            wb_tpl = openpyxl.load_workbook(template_path, data_only=True)
            ws_tpl = wb_tpl.active
            for row in ws_tpl.iter_rows(min_row=1, max_row=50, max_col=10, values_only=False):
                for cell in row:
                    if cell.value is not None:
                        template_cells[(cell.row, cell.column)] = str(cell.value).strip()
            wb_tpl.close()
        except Exception as e:
            logger.warning(f'[验证] 无法加载原模板用于对比: {e}')

    spec_col = layout.data_columns.get('spec', 3)
    result_col = layout.data_columns.get('result', 5)
    method_col = layout.data_columns.get('method', 6)

    details = []
    template_defaults_retained = []

    def _check_template_default(field_name: str, row: int, col: int, actual: str, expected_from_pdf: str):
        """检测是否保留了模板默认值"""
        if not template_cells or not actual:
            return
        tpl_val = template_cells.get((row, col), "")
        if tpl_val and actual == tpl_val and expected_from_pdf and \
           _normalize_for_compare(actual) != _normalize_for_compare(expected_from_pdf):
            template_defaults_retained.append({
                "field": field_name,
                "template_default": tpl_val,
                "pdf_value": expected_from_pdf,
                "output_value": actual,
            })
            logger.warning(f'[验证-默认值残留] {field_name}: 输出={actual!r} 与模板默认值相同，'
                           f'但 PDF 数据为 {expected_from_pdf!r}')

    # --- 验证头部字段 ---
    for std_key, mapping in layout.header_fields.items():
        expected = coa.header.get(std_key, "")
        if std_key in ("mfg_date", "exp_date") and expected:
            expected = convert_date(expected)

        cell_ref = mapping.value_pos.cell_ref
        if cell_ref:
            actual = ws[cell_ref].value
        else:
            actual = ws.cell(row=mapping.value_pos.row, column=mapping.value_pos.col).value
        actual = str(actual) if actual is not None else ""

        status = _compare_values(expected, actual)
        details.append({
            "field": f"header.{std_key}",
            "expected": expected,
            "actual": actual,
            "status": status,
        })
        _check_template_default(f"header.{std_key}",
                                mapping.value_pos.row, mapping.value_pos.col, actual, expected)

    # --- 验证 Assay 行 ---
    assay_row = layout.table_rows.get('assay')
    if assay_row:
        if coa.assay:
            for col_key, col_num in [('specification', spec_col), ('result', result_col), ('method', method_col)]:
                expected = coa.assay.get(col_key, "")
                actual = ws.cell(row=assay_row, column=col_num).value
                actual = str(actual) if actual is not None else ""
                status = _compare_values(expected, actual)
                details.append({
                    "field": f"assay.{col_key}",
                    "expected": expected,
                    "actual": actual,
                    "status": status,
                })
        else:
            for col_key, col_num in [('specification', spec_col), ('result', result_col), ('method', method_col)]:
                actual = ws.cell(row=assay_row, column=col_num).value
                actual = str(actual) if actual is not None else ""
                status = "pass" if not actual else "fail"
                details.append({
                    "field": f"assay.{col_key}",
                    "expected": "",
                    "actual": actual,
                    "status": status,
                })

    # --- 验证数据行 ---
    extracted = {}
    for item in coa.analytical_items + coa.microbiology_items:
        key = normalize_item_name(item["item"])
        if key:
            extracted[key] = item

    for item_key, row_num in layout.table_rows.items():
        if item_key == 'assay':
            continue

        if item_key in extracted:
            item = extracted[item_key]
            for col_key, col_num in [('specification', spec_col), ('result', result_col), ('method', method_col)]:
                expected = item.get(col_key, "")
                actual = ws.cell(row=row_num, column=col_num).value
                actual = str(actual) if actual is not None else ""
                status = _compare_values(expected, actual)
                details.append({
                    "field": f"data.{item_key}.{col_key}",
                    "expected": expected,
                    "actual": actual,
                    "status": status,
                })
                _check_template_default(f"data.{item_key}.{col_key}", row_num, col_num, actual, expected)
        else:
            for col_key, col_num in [('specification', spec_col), ('result', result_col), ('method', method_col)]:
                actual = ws.cell(row=row_num, column=col_num).value
                actual = str(actual) if actual is not None else ""
                status = "pass" if not actual else "fail"
                details.append({
                    "field": f"data.{item_key}.{col_key}",
                    "expected": "(空-PDF无此项)",
                    "actual": actual,
                    "status": status,
                })

    # --- 验证 Packing & Storage ---
    if layout.packing_position:
        if coa.packing_storage:
            parts = re.split(r'(?:store\s)', coa.packing_storage, maxsplit=1, flags=re.IGNORECASE)
            expected_packing = parts[0].strip().rstrip('.') if parts else ""
        else:
            expected_packing = ""

        ref = layout.packing_position.cell_ref
        p_row, p_col = layout.packing_position.row, layout.packing_position.col
        if ref:
            actual = ws[ref].value
        else:
            actual = ws.cell(row=p_row, column=p_col).value
        actual = str(actual) if actual is not None else ""

        # 对 packing 做模板默认值检测（即使 PDF 期望值为空）
        tpl_packing = template_cells.get((p_row, p_col), "") if template_cells else ""
        if not expected_packing and actual and tpl_packing and actual == tpl_packing:
            # PDF 没有提取到 packing，但输出保留了模板默认值 → 标记为 fail
            status = "fail"
            template_defaults_retained.append({
                "field": "packing",
                "template_default": tpl_packing,
                "pdf_value": "(PDF未提取到)",
                "output_value": actual,
            })
            logger.warning(f'[验证-默认值残留] packing: 输出保留模板默认值 {actual!r}，PDF 未提取到对应数据')
        else:
            status = _compare_values(expected_packing, actual)
            _check_template_default("packing", p_row, p_col, actual, expected_packing)

        details.append({
            "field": "packing",
            "expected": expected_packing or "(空-PDF无此项)",
            "actual": actual,
            "status": status,
        })

    if layout.storage_position:
        if coa.packing_storage:
            parts = re.split(r'(?:store\s)', coa.packing_storage, maxsplit=1, flags=re.IGNORECASE)
            expected_storage = ("Store " + parts[1].strip()) if len(parts) >= 2 else ""
        else:
            expected_storage = ""

        ref = layout.storage_position.cell_ref
        s_row, s_col = layout.storage_position.row, layout.storage_position.col
        if ref:
            actual = ws[ref].value
        else:
            actual = ws.cell(row=s_row, column=s_col).value
        actual = str(actual) if actual is not None else ""

        # 对 storage 做模板默认值检测
        tpl_storage = template_cells.get((s_row, s_col), "") if template_cells else ""
        if not expected_storage and actual and tpl_storage and actual == tpl_storage:
            status = "fail"
            template_defaults_retained.append({
                "field": "storage",
                "template_default": tpl_storage,
                "pdf_value": "(PDF未提取到)",
                "output_value": actual,
            })
            logger.warning(f'[验证-默认值残留] storage: 输出保留模板默认值 {actual!r}，PDF 未提取到对应数据')
        else:
            status = _compare_values(expected_storage, actual)
            _check_template_default("storage", s_row, s_col, actual, expected_storage)

        details.append({
            "field": "storage",
            "expected": expected_storage or "(空-PDF无此项)",
            "actual": actual,
            "status": status,
        })

    wb.close()

    # --- 汇总 ---
    total = len(details)
    passed = sum(1 for d in details if d["status"] in ("pass", "empty_ok"))
    failed = sum(1 for d in details if d["status"] == "fail")
    accuracy = passed / total if total > 0 else 1.0

    result = {
        "total": total,
        "passed": passed,
        "failed": failed,
        "accuracy": accuracy,
        "details": details,
        "template_defaults_retained": template_defaults_retained,
    }

    # 输出验证报告日志
    logger.info(f'[验证] 总字段: {total}, 通过: {passed}, 失败: {failed}, 正确率: {accuracy:.1%}')
    if template_defaults_retained:
        logger.warning(f'[验证] 发现 {len(template_defaults_retained)} 个字段保留了模板默认值')
    for d in details:
        if d["status"] == "fail":
            logger.error(f'[验证-失败] {d["field"]}: 期望={d["expected"]!r}, 实际={d["actual"]!r}')

    return result


def _compare_values(expected: str, actual: str) -> str:
    """比较期望值和实际值，返回 pass/fail/empty_ok"""
    expected = (expected or "").strip()
    actual = (actual or "").strip()

    if not expected and not actual:
        return "empty_ok"
    if expected == actual:
        return "pass"
    # 宽松比较：忽略多余空格和大小写差异
    if _normalize_for_compare(expected) == _normalize_for_compare(actual):
        return "pass"
    return "fail"


def _normalize_for_compare(s: str) -> str:
    """标准化字符串用于宽松比较"""
    s = re.sub(r'\s+', ' ', s).strip().lower()
    # 移除尾部句号
    s = s.rstrip('.')
    return s


